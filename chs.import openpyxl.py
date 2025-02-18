import openpyxl
import pyperclip
import os

# 엑셀 파일이 있는 폴더 경로
directory_path = r"C:\Users\jaguar\Downloads"  # 원하는 폴더 경로로 변경

# 파일 선택 프로세스
while True:
    file_name = input(f"읽을 엑셀 파일명을 입력하세요 (확장자 생략 가능, 예: data): ").strip()
    
    # 확장자가 없으면 .xlsx 추가
    if not file_name.lower().endswith(".xlsx"):
        file_name += ".xlsx"

    file_path = os.path.join(directory_path, file_name)
    
    if os.path.exists(file_path):
        break
    else:
        print("❌ 파일이 존재하지 않습니다. 다시 입력해주세요.")

# 엑셀 파일 열기
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active  # 첫 번째 시트 활성화

# 찾을 키워드
keywords = ["중간_CNS", "zh-hans", "CNS", "zh_CN"]

# 특정 키워드가 있는 행과 열 찾기
target_row = None
target_column = None

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
    for col_idx, cell_value in enumerate(row, start=1):
        if cell_value in keywords:  # 키워드가 포함된 행 찾기
            target_row = row_idx
            target_column = col_idx
            break
    if target_row:
        break  # 첫 번째 일치하는 키워드만 찾음

# 키워드가 발견되지 않으면 종료
if target_row is None:
    print("❌ 키워드를 찾을 수 없습니다.")
else:
    # 키워드 행 바로 아래부터 끝까지 해당 열의 데이터 가져오기
    values = [
        str(ws.cell(row=i, column=target_column).value).replace("\n", "\r\n")  # 줄바꿈 유지
        for i in range(target_row + 1, ws.max_row + 1)
        if ws.cell(row=i, column=target_column).value is not None
    ]

    # 엑셀에서 붙여넣을 때 한 셀 안에 줄바꿈이 유지되도록 " "로 감싸기
    if values:
        formatted_text = "\r\n".join(f'"{value}"' for value in values)  # 각 셀 값을 " "로 감싸기
        pyperclip.copy(formatted_text)
        print(f"✅ 클립보드에 복사 완료!\n{formatted_text}")
    else:
        print("⚠️ 복사할 데이터가 없습니다.")

# 워크북 닫기
wb.close()
