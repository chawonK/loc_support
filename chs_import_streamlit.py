import openpyxl
import pyperclip
import os
import streamlit as st

# 웹 페이지 제목
st.title('엑셀 데이터 복사')
st.caption(":rainbow[지정된 키워드 바로 아래 행부터 전체 내용이 클립보드에 복사됩니다.]")


# 미리 설정된 폴더 경로 및 키워드
default_directory_path = "C:/Users/jaguar/Downloads"  # 기본 경로
default_keywords = ["중간_CNS", "zh-hans", "CNS", "zh_CN","Simplified Chinese"]  # 기본 키워드

# 폴더 경로 입력 (사용자가 수정할 수 있음)
directory_path = st.text_input("파일이 있는 폴더 경로", value=default_directory_path)

# 🚀 **디버깅용 출력 (Streamlit에서 직접 확인)**
st.write(f"🔍 입력된 폴더 경로: `{directory_path}`")
st.write(f"✅ 경로 존재 여부: `{os.path.exists(directory_path)}`")

# ✅ 폴더 존재 여부 확인
if not os.path.isdir(directory_path):  # `isdir()` 사용
    st.error("❌ 입력한 폴더 경로가 존재하지 않습니다. 올바른 경로를 입력하세요.")
    xlsx_files = []
else:
    # 📄 폴더 내 엑셀 파일 목록 가져오기 (수정된 날짜 기준 정렬)
    xlsx_files = sorted(
        [f for f in os.listdir(directory_path) if f.endswith(".xlsx")],
        key=lambda f: os.path.getmtime(os.path.join(directory_path, f)), 
        reverse=True  # 최신 파일이 제일 위로 오도록 정렬
    )

# 파일 선택 (자동으로 불러오기)
if xlsx_files:
    file_name = st.selectbox("📄 파일 선택", xlsx_files)
else:
    file_name = None
    st.warning("⚠️ 해당 폴더에 `.xlsx` 파일이 없습니다. 경로를 확인하세요.")

# 키워드 입력 (사용자가 수정할 수 있음)
keywords_input = st.text_area("찾을 키워드(언어열 이름)", value=", ".join(default_keywords))

# 버튼 클릭 시 실행
if st.button("실행"):
    # 입력값이 모두 있는지 확인
    if not file_name:
        st.error("파일명을 입력해주세요!")
    else:
        # 키워드 처리: 쉼표로 구분된 키워드 리스트로 변환
        keywords = [keyword.strip() for keyword in keywords_input.split(',')]

        # 파일명에 확장자 .xlsx가 없다면 자동으로 추가
        if not file_name.lower().endswith(".xlsx"):
            file_name += ".xlsx"

        file_path = os.path.join(directory_path, file_name)

        if os.path.exists(file_path):
            # 엑셀 파일 열기
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active  # 첫 번째 시트 활성화

            # 특정 키워드가 있는 행과 열 찾기
            target_row = None
            target_column = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value in keywords:  # 키워드가 포함된 행 찾기
                        target_row = row_idx
                        target_column = col_idx
                        break
                if target_row:
                    break  # 첫 번째 일치하는 키워드만 찾음

            # 키워드가 발견되지 않으면 종료
            if target_row is None:
                st.error("❌ 키워드를 찾을 수 없습니다.")
            else:
                # 키워드 행 바로 아래부터 끝까지 해당 열의 데이터 가져오기
                values = [
                    str(ws.cell(row=i, column=target_column).value).replace("\n", "\r\n")  # 줄바꿈 유지
                    for i in range(target_row + 1, ws.max_row + 1)
                    if ws.cell(row=i, column=target_column).value is not None
                ]
                
                # 엑셀에서 붙여넣을 때 한 셀 안에 줄바꿈이 유지되도록 " "로 감싸기
                if values:
                    formatted_text = "\r\n".join(f'"{value}"' for value in values)  # 각 셀 값을 " "로 감싸기
                    pyperclip.copy(formatted_text)
                    
                    # 성공 메시지와 formatted_text를 다른 영역에 표시
                    st.success("✅ 클립보드에 복사 완료!")
                    st.text_area("복사된 내용", formatted_text, height=200)
                else:
                    st.warning("⚠️ 복사할 데이터가 없습니다.")
            
            # 워크북 닫기
            wb.close()
        else:
            st.error("❌ 파일이 존재하지 않습니다.")
