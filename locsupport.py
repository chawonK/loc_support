import streamlit as st
import streamlit.components.v1 as components
import openpyxl
import pandas as pd
import re
import zipfile
from io import BytesIO

st.set_page_config(page_title="엑셀 도구 모음", layout="centered")

# 사이드바 메뉴
st.sidebar.title("엑셀 도구 모음")
page = st.sidebar.radio(" ", ("엑셀 데이터 복사 & 미리보기", "엑셀 시트 분할", "엑셀 분할(분류별)", "단어수 카운터(웹)", "Vlookup 매칭", "월간 보고 데이터"))

# 1. 엑셀 데이터 복사 + 미리보기
if page == "엑셀 데이터 복사 & 미리보기":
    st.title("📄 엑셀 데이터 복사 & 미리보기")
    st.write(":rainbow[지정된 키워드 바로 아래 행부터 복사하며, 시트 내용을 함께 미리볼 수 있습니다.]")

    # 파일 업로드
    uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsx", "xls"])

    default_keywords = ["중간_CNS", "중간_CHS", "zh-hans", "CNS", "CHS", "zh_CN", "Simplified Chinese", "CNS (중국어 간체)"]
    keywords_input = st.text_area("찾을 키워드(언어열 이름)", value=", ".join(default_keywords))

    if uploaded_file:
        # 시트 선택
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names
        selected_sheet = st.selectbox("시트를 선택하세요", sheet_names)

        # 복사 처리 (openpyxl)
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws = wb[selected_sheet]
        keywords = [k.strip() for k in keywords_input.split(',')]

        target_row, target_column = None, None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value in keywords:
                    target_row, target_column = row_idx, col_idx
                    break
            if target_row:
                break

        if target_row is None:
            st.error("❌ 키워드를 찾을 수 없습니다.")
        else:
            values = [str(ws.cell(row=i, column=target_column).value or "") for i in range(target_row + 1, ws.max_row + 1)]
            formatted_text = "\n".join(f'"{value}"' for value in values)

            if values:
                st.success("✅ 데이터 추출 완료! 아래에서 복사하거나 다운로드하세요.")
                st.text_area("복사할 텍스트", formatted_text, height=200)

                # JS 복사 버튼 (Streamlit Components)
                js_text = formatted_text.replace("\n", "\\n").replace('"', '\\"')
                components.html(f"""
                    <button onclick="navigator.clipboard.writeText('{js_text}')">
                        📋 복사하기 (클립보드)
                    </button>
                """, height=40)

                # 텍스트 파일 다운로드
                st.download_button("📥 텍스트 다운로드", formatted_text, "extracted_data.txt")
            else:
                st.warning("⚠️ 복사할 데이터가 없습니다.")

        wb.close()

        # 시트 미리보기
        df = pd.read_excel(xls, sheet_name=selected_sheet)
        if not df.empty:
            st.divider()
            st.subheader(f"🔍 '{selected_sheet}' 시트 미리보기")
            st.dataframe(df.head(20))
        else:
            st.warning(f"⚠️ '{selected_sheet}' 시트에 표시할 데이터가 없습니다.")
    else:
        st.caption("파일을 업로드해 주세요.")


# 2. 엑셀 시트 분할
elif page == "엑셀 시트 분할":
    st.title("✂️ 엑셀 시트 분할")
    st.caption("※ 엑셀 파일의 각 시트를 새로운 파일로 분할하여 ZIP으로 다운로드합니다.")
    
    uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요 (.xlsx)", type=["xlsx"])
    
    if uploaded_file:
        file_name = uploaded_file.name
        excel_file = pd.ExcelFile(uploaded_file)
    
        if not excel_file.sheet_names:
            st.error("❌ 해당 엑셀 파일에 시트가 없습니다.")
        else:
            # ZIP 파일을 메모리에 생성
            zip_buffer = BytesIO()
            zip_file = zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED)
    
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name)
                
                # 각 시트를 엑셀 파일로 변환
                sheet_io = BytesIO()
                with pd.ExcelWriter(sheet_io, engine="xlsxwriter") as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # ZIP에 추가
                zip_file.writestr(f"{sheet_name}.xlsx", sheet_io.getvalue())
    
            zip_file.close()  # ZIP 파일 닫기
    
            # ZIP 파일 다운로드
            zip_buffer.seek(0)
            st.download_button(
                label="📥 ZIP 파일 다운로드",
                data=zip_buffer,
                file_name=f"{file_name}_시트분할.zip",
                mime="application/zip",
            )

# 2-2. 엑셀 분할(분류별)
elif page == "엑셀 분할(분류별)":
    st.title("✂️ 엑셀 분할(분류별)")
    st.caption("※ 엑셀 파일의 특정 열을 기준으로 데이터를 분할하여 각 분류별로 개별 엑셀 파일로 저장합니다.")
    
    uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요 (.xlsx)", type=["xlsx"])

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
    
        # 사용자에게 분류 기준 열 선택 UI 제공
        columns = df.columns.tolist()
        selected_column = st.selectbox("분류 기준 열을 선택하세요", columns)
    
        if st.button("🚀 분할 실행 (ZIP 파일 다운로드)", use_container_width=True):
            grouped = df.groupby(selected_column)
    
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for category, group in grouped:
                    sanitized_name = re.sub(r'[\\/*?:"<>|]', "_", str(category))
                    excel_buffer = BytesIO()
                    group.to_excel(excel_buffer, index=False)
                    excel_buffer.seek(0)
                    zip_file.writestr(f"{sanitized_name}.xlsx", excel_buffer.read())
    
            zip_buffer.seek(0)
            st.success("✅ 분류별 엑셀 파일 분할 완료! ZIP 다운로드 가능")
            st.download_button(
                label="📦 ZIP 파일 다운로드",
                data=zip_buffer,
                file_name="분류별_엑셀_분할.zip",
                mime="application/zip"
            )

# 3. 월간 보고 데이터
elif page == "월간 보고 데이터":
    st.title("📊 Jira CSV 데이터 추출기")
    st.caption("Jira에서 요청 필터 페이지에서 우측 상단의 '보기' > 'CSV (모든 필드)'로 다운로드 받은 월별 전체 요청 파일을 업로드 하면 월별 프로젝트별 요청수, 단어수 합계를 볼 수 있습니다.")
    uploaded_file = st.file_uploader("CSV 파일 업로드", type=["csv"])

    if uploaded_file:
        df = pd.read_csv(uploaded_file)
        df_filtered = df[['프로젝트 이름', '요약', '기한', '생성일']].copy()
        df_filtered[['단어수', '기준 언어']] = df_filtered['요약'].str.extract(r'\[(\d+)\s*([A-Za-z]+)\]')
        df_filtered['단어수'] = pd.to_numeric(df_filtered['단어수'], errors='coerce')
        project_summary_df = df_filtered.groupby('프로젝트 이름').agg(요청수=('요약', 'count'), 단어수_합계=('단어수', 'sum')).reset_index()
        total_row = pd.DataFrame({'프로젝트 이름': ['합계'], '요청수': [project_summary_df['요청수'].sum()], '단어수_합계': [project_summary_df['단어수_합계'].sum()]})
        project_summary_df = pd.concat([project_summary_df, total_row], ignore_index=True)

        # C2 셀에서 시트명을 동적으로 생성
        c2_value = df_filtered['기한'].iloc[0]  # C2 내용 (기한 열의 첫 번째 값)
        sheet_name_prefix = c2_value[2:4] + c2_value[5:7]  # 2, 3, 5, 6번째 글자 추출
        original_sheet_name = sheet_name_prefix  # '원본 데이터' 시트명
        summary_sheet_name = sheet_name_prefix + " 월별 통계"  # '프로젝트별 요약' 시트명

        # 동적으로 파일명 생성
        file_name = f"{sheet_name_prefix}_project_summary.xlsx"

        # Excel 파일로 저장
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_filtered.to_excel(writer, index=False, sheet_name=original_sheet_name)
            project_summary_df.to_excel(writer, index=False, sheet_name=summary_sheet_name)
        output.seek(0)

        # 다운로드 버튼
        st.download_button(
            label="📥 엑셀 파일 다운로드",
            data=output,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# 4. 단어수 카운터
elif page == "단어수 카운터(웹)":
    st.title("🔢 단어수 카운터(웹)")
    st.write("텍스트를 입력하면 띄어쓰기 기준으로 단어 수를 계산합니다.")

    def count_words(text):
        words = text.split()
        return len(words)

    if 'word_count' not in st.session_state:
        st.session_state.word_count = 0

    st.subheader(f"단어 수: {st.session_state.word_count}")

    def update_word_count():
        st.session_state.word_count = count_words(st.session_state.text_input)

    text_input = st.text_area("텍스트 입력", height=200, key='text_input', on_change=update_word_count)

# 5. VLOOKUP 스타일 파일 매칭기
elif page == "Vlookup 매칭":
    st.title("🔎 VLOOKUP 스타일 파일 매칭기")
    st.write("기준 파일과 내용을 가져올 파일을 업로드 하면 매칭할 열을 선택할 수 있습니다.")

    def vlookup_page():
        # 파일 업로드
        base_file = st.file_uploader("기준 파일을 업로드하세요", type=["xlsx", "csv"])
        lookup_file = st.file_uploader("내용을 가져올 파일을 업로드하세요", type=["xlsx", "csv"])

        if base_file and lookup_file:
            # 파일 읽기 with 예외처리
            try:
                if base_file.name.endswith('csv'):
                    df_base = pd.read_csv(base_file)
                else:
                    df_base = pd.read_excel(base_file)
            except Exception as e:
                st.error(f"기준 파일 읽기 실패: {e}")
                st.stop()

            try:
                if lookup_file.name.endswith('csv'):
                    df_lookup = pd.read_csv(lookup_file)
                else:
                    df_lookup = pd.read_excel(lookup_file)
            except Exception as e:
                st.error(f"내용 파일 읽기 실패: {e}")
                st.stop()

            # 데이터 비어있는지 체크
            if df_base.empty or df_lookup.empty:
                st.error("업로드된 파일에 데이터가 없습니다. 다시 확인해주세요.")
                st.stop()

            st.subheader("기준 파일 열 선택")
            key_base = st.selectbox("매칭에 사용할 기준 파일 열 선택", df_base.columns)

            st.subheader("내용을 가져올 파일 열 선택")
            key_lookup = st.selectbox("매칭에 사용할 내용 가져올 파일 열 선택", df_lookup.columns)

            st.subheader("가져올 내용의 열 (여러 개 선택 가능)")
            value_columns = st.multiselect("내용 가져올 파일에서 가져올 열을 선택하세요", df_lookup.columns)

            # 매칭 시작
            if st.button("매칭 시작"):
                df_result = df_base.copy()
                for col in value_columns:
                    lookup_dict = dict(zip(df_lookup[key_lookup], df_lookup[col]))
                    df_result[f"{col}_가져옴"] = df_result[key_base].map(lookup_dict)

                st.success("매칭 완료!")
                st.dataframe(df_result)

                # 결과 다운로드
                @st.cache_data
                def convert_df(df):
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    output.seek(0)
                    return output.getvalue()

                excel_data = convert_df(df_result)
                st.download_button(
                    label="📥 결과 파일 다운로드",
                    data=excel_data,
                    file_name='merged_result.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    # 함수 실행
    vlookup_page()

